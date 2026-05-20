"""
Generate mock test invoices for Maverick upload flow demos.
Creates 2 PDFs (PCD drayage + CDS drayage) and 1 Excel — all with
brand-new invoice numbers so they bubble to the top of the table
(JUST_UPLOADED set) and don't collide with seed data.

Run from inside maverick_backend/:
    python3 test_invoices/_generate.py
"""
import os
from reportlab.lib.pagesizes import LETTER
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether,
)
from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))


def _styles():
    s = getSampleStyleSheet()
    s.add(ParagraphStyle(name="H1Big", fontName="Helvetica-Bold", fontSize=14, spaceAfter=6))
    s.add(ParagraphStyle(name="Small", fontName="Helvetica", fontSize=8, leading=10))
    s.add(ParagraphStyle(name="SmallB", fontName="Helvetica-Bold", fontSize=8, leading=10))
    return s


def _build_pdf(path, header_company, header_addr, header_contact, invoice_no,
               invoice_date, terms, fb_no, container_no, bol, origin, dest, equip,
               linehaul, fsc_pct, fsc_amount, accessorials, grand_total, msa_ref):
    doc = SimpleDocTemplate(
        path, pagesize=LETTER,
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
        topMargin=0.5 * inch, bottomMargin=0.5 * inch,
    )
    s = _styles()
    story = []

    story.append(Paragraph(f"<b>{header_company}</b>  ·  Invoice {invoice_no}", s["Small"]))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>{header_company}</b>", s["H1Big"]))

    meta = [
        ["INVOICE #", invoice_no, "Invoice Date", invoice_date],
        ["", "", "Shipment Date", invoice_date],
        ["", "", "Terms", terms],
    ]
    t = Table(meta, colWidths=[1.1 * inch, 2.2 * inch, 1.1 * inch, 2.2 * inch])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.grey),
        ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(t)
    story.append(Spacer(1, 6))
    story.append(Paragraph(header_addr, s["Small"]))
    story.append(Paragraph(header_contact, s["Small"]))
    story.append(Spacer(1, 8))

    story.append(Paragraph("<b>BILL TO</b>", s["SmallB"]))
    story.append(Paragraph("NewAge Products Logistics California Inc.", s["Small"]))
    story.append(Paragraph("3125 Wilson Ave, Perris, CA 92571", s["Small"]))
    story.append(Paragraph("AP: ap.usa@newageproducts.com · ATTN: Inbound Freight", s["Small"]))
    story.append(Spacer(1, 8))

    story.append(Paragraph("<b>LOAD DETAILS</b>", s["SmallB"]))
    story.append(Paragraph(f"FB# / Load ID: {fb_no}", s["Small"]))
    story.append(Paragraph(f"Container #: {container_no}", s["Small"]))
    story.append(Paragraph(f"BOL / MBL: {bol}", s["Small"]))
    story.append(Paragraph(f"Equipment: {equip}", s["Small"]))
    story.append(Paragraph(f"Origin: {origin}", s["Small"]))
    story.append(Paragraph(f"Destination: {dest}", s["Small"]))
    story.append(Spacer(1, 8))

    # Line table — extractor finds the table whose first header cell is "#"
    fsc_amt_str = f"{fsc_amount:.2f}"
    linehaul_str = f"{linehaul:.2f}"
    shipment_amt = round(linehaul + fsc_amount, 2)
    rows = [
        ["#", "Type", "Description", "Qty", "Rate (USD)", "Linehaul\n(USD)", "FSC %", "FSC\n(USD)", "Amount\n(USD)"],
        ["1", "Shipment", "Linehaul (P&D) + Fuel Surcharge", "1",
         linehaul_str, linehaul_str, f"{int(fsc_pct * 100)}%", fsc_amt_str, f"{shipment_amt:.2f}"],
    ]
    n = 2
    for desc, amt in accessorials:
        rows.append([str(n), "Accessorial", desc, "1", f"{amt:.2f}", "—", "—", "—", f"{amt:.2f}"])
        n += 1
    line_tbl = Table(rows, colWidths=[0.3 * inch, 0.75 * inch, 2.0 * inch, 0.4 * inch,
                                      0.75 * inch, 0.75 * inch, 0.45 * inch, 0.6 * inch, 0.75 * inch])
    line_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eaeaea")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(line_tbl)
    story.append(Spacer(1, 6))

    accessorials_sum = sum(a[1] for a in accessorials)
    summary = [
        ["Linehaul subtotal", f"${linehaul:.2f}"],
        ["FSC subtotal", f"${fsc_amount:.2f}"],
        ["Accessorials subtotal", f"${accessorials_sum:.2f}"],
        ["GRAND TOTAL (USD)", f"${grand_total:.2f}"],
    ]
    sum_tbl = Table(summary, colWidths=[2.5 * inch, 1.2 * inch], hAlign="RIGHT")
    sum_tbl.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.75, colors.black),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 6))
    story.append(Paragraph("(US drayage — no sales tax on freight services)", s["Small"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph("<b>REMIT TO</b>", s["SmallB"]))
    story.append(Paragraph(f"{header_company} · Wells Fargo Bank · Acct ending 7401 · Routing 121000248", s["Small"]))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"Disputes within 30 days per MSA {msa_ref} §7.2. Late payments accrue 1.5% per month.",
        s["Small"]))

    doc.build(story)
    print(f"  wrote {os.path.basename(path)}")


def _build_customs_pdf(path, invoice_no, invoice_date, period, terms, entries, grand_total):
    """Customs broker invoice (Livingston-style). PDF text includes entry #s,
    container links, HTS, duty stack — same extractor reads it; backend audit
    fires when grand_total > $50k."""
    doc = SimpleDocTemplate(
        path, pagesize=LETTER,
        leftMargin=0.45 * inch, rightMargin=0.45 * inch,
        topMargin=0.5 * inch, bottomMargin=0.5 * inch,
    )
    s = _styles()
    story = []
    header = "LIVINGSTON INTERNATIONAL"
    story.append(Paragraph(f"<b>{header}</b>  ·  Invoice {invoice_no}", s["Small"]))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>{header} — U.S. Broker Invoice</b>", s["H1Big"]))

    meta = [
        ["INVOICE #", invoice_no, "Invoice Date", invoice_date],
        ["", "", "Period", period],
        ["", "", "Terms", terms],
    ]
    t = Table(meta, colWidths=[1.1 * inch, 2.2 * inch, 1.1 * inch, 2.5 * inch])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.grey),
        ("LINEBELOW", (0, -1), (-1, -1), 0.5, colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(t)
    story.append(Spacer(1, 6))
    story.append(Paragraph("1140 Spruce St, Trenton, NJ 08648 · AP@livingstonintl.com · Filer Code 8869", s["Small"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph("<b>BILL TO / IMPORTER OF RECORD</b>", s["SmallB"]))
    story.append(Paragraph("NewAge Products USA, Inc. · EIN 47-1820392", s["Small"]))
    story.append(Paragraph("3125 Wilson Ave, Perris, CA 92571 · AP: ap.usa@newageproducts.com", s["Small"]))
    story.append(Spacer(1, 10))

    # Entries table — header starts with "#" so pdfplumber.extract_tables picks it up
    rows = [["#", "Entry #", "Container", "PO", "Entered Value", "HTS Code",
             "Duty Rate", "Sec 301", "Sec 232", "Duty", "MPF", "HMF", "Brokerage", "Subtotal"]]
    for i, e in enumerate(entries, 1):
        subtotal = round(e["duty"] + e["mpf"] + e["hmf"] + e["brokerage"] + e["disbursement"] + e["isf"], 2)
        rows.append([
            str(i), e["entry"], e["container"], e["po"],
            f"${e['value']:.2f}", e["hts"], e["duty_rate"], e["sec301"], e["sec232"],
            f"${e['duty']:.2f}", f"${e['mpf']:.2f}", f"${e['hmf']:.2f}",
            f"${e['brokerage']:.2f}", f"${subtotal:.2f}",
        ])
    line_tbl = Table(rows, colWidths=[0.25, 0.6, 0.7, 0.7, 0.7, 0.65, 0.5, 0.4, 0.4, 0.6, 0.45, 0.45, 0.55, 0.65])
    # convert to inch-friendly widths
    line_tbl._argW = [w * inch for w in line_tbl._argW]
    line_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eaeaea")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(line_tbl)
    story.append(Spacer(1, 8))

    duty_sum = sum(e["duty"] for e in entries)
    summary = [
        ["Total Duty", f"${duty_sum:.2f}"],
        ["MPF + HMF", f"${sum(e['mpf'] + e['hmf'] for e in entries):.2f}"],
        ["Brokerage + disbursement + ISF", f"${sum(e['brokerage'] + e['disbursement'] + e['isf'] for e in entries):.2f}"],
        ["GRAND TOTAL (USD)", f"${grand_total:.2f}"],
    ]
    sum_tbl = Table(summary, colWidths=[2.8 * inch, 1.2 * inch], hAlign="RIGHT")
    sum_tbl.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.75, colors.black),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "Per CBP Form 7501. Section 301 (China-origin) and Section 232 stack atop HTS base rate. "
        "Disputes within 30 days per broker MSA §6.4.", s["Small"]))

    doc.build(story)
    print(f"  wrote {os.path.basename(path)}")


def _build_customs_excel(path, broker_invoice_no, invoice_date, period, terms, entries):
    """Customs broker invoice in Excel form — 10 entries on one sheet.
    Mirrors the master-workbook customs schema so seed.py / extractor logic
    can read it. Filename contains 'customs' so backend api_upload_customs
    flags it as customs and runs the duty-math audit."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Customs_Invoice"

    # Header block — broker invoice meta
    ws["A1"] = "LIVINGSTON INTERNATIONAL — U.S. Broker Invoice"
    ws["A2"] = "Invoice #"
    ws["B2"] = broker_invoice_no
    ws["A3"] = "Invoice Date"
    ws["B3"] = invoice_date
    ws["A4"] = "Period"
    ws["B4"] = period
    ws["A5"] = "Bill-To"
    ws["B5"] = "NewAge Products USA, Inc."
    ws["A6"] = "Importer of Record"
    ws["B6"] = "NewAge Products USA, Inc. · EIN 47-1820392"
    ws["A7"] = "Terms"
    ws["B7"] = terms

    # Entries table starts row 9
    headers = [
        "Entry #", "Container #", "PO #", "Entered Value (USD)", "HTS Code",
        "Duty Rate", "Sec 301 %", "Sec 232 %", "Duty (USD)", "MPF (USD)",
        "HMF (USD)", "Brokerage (USD)", "Disbursement (USD)", "ISF (USD)",
        "Subtotal (USD)", "Notes",
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=9, column=col_idx, value=h)

    grand_total = 0.0
    for i, e in enumerate(entries, 10):
        subtotal = round(
            e["duty"] + e["mpf"] + e["hmf"]
            + e["brokerage"] + e["disbursement"] + e["isf"], 2
        )
        grand_total += subtotal
        ws.cell(row=i, column=1, value=e["entry"])
        ws.cell(row=i, column=2, value=e["container"])
        ws.cell(row=i, column=3, value=e["po"])
        ws.cell(row=i, column=4, value=e["value"])
        ws.cell(row=i, column=5, value=e["hts"])
        ws.cell(row=i, column=6, value=e["duty_rate"])
        ws.cell(row=i, column=7, value=e["sec301"])
        ws.cell(row=i, column=8, value=e["sec232"])
        ws.cell(row=i, column=9, value=e["duty"])
        ws.cell(row=i, column=10, value=e["mpf"])
        ws.cell(row=i, column=11, value=e["hmf"])
        ws.cell(row=i, column=12, value=e["brokerage"])
        ws.cell(row=i, column=13, value=e["disbursement"])
        ws.cell(row=i, column=14, value=e["isf"])
        ws.cell(row=i, column=15, value=subtotal)
        ws.cell(row=i, column=16, value=e.get("notes", ""))

    # Grand-total footer
    foot_row = 10 + len(entries) + 1
    ws.cell(row=foot_row, column=14, value="GRAND TOTAL (USD)")
    ws.cell(row=foot_row, column=15, value=round(grand_total, 2))

    # "Invoices" + "Invoice_Lines" sheets — matches Maverick's multi-sheet
    # extractor (headers on ROW 4, data starts ROW 5). Lets the backend
    # extractor pull grand_total + carrier + invoice meta cleanly so the
    # duty-math audit can fire on grand_total > $50k.
    ws2 = wb.create_sheet("Invoices")
    ws2["A1"] = "LIVINGSTON INTERNATIONAL — Customs Broker Invoice"
    ws2["A2"] = f"Period: {period}"
    inv_headers = [
        "#", "Invoice #", "Carrier", "Invoice Date", "FB# / Load ID",
        "Container #", "BOL/MBL #", "Origin", "Destination", "Rate Type",
        "Linehaul (USD)", "FSC %", "FSC (USD)", "Accessorials (USD)",
        "Grand Total (USD)",
    ]
    for col_idx, h in enumerate(inv_headers, 1):
        ws2.cell(row=4, column=col_idx, value=h)
    ws2.cell(row=5, column=1, value=1)
    ws2.cell(row=5, column=2, value=broker_invoice_no)
    ws2.cell(row=5, column=3, value="Livingston International")
    ws2.cell(row=5, column=4, value=invoice_date)
    ws2.cell(row=5, column=5, value="—")
    ws2.cell(row=5, column=6, value="MULTI")
    ws2.cell(row=5, column=7, value="—")
    ws2.cell(row=5, column=8, value="Multiple POEs")
    ws2.cell(row=5, column=9, value="NewAge DCs")
    ws2.cell(row=5, column=10, value="CUSTOMS")
    ws2.cell(row=5, column=11, value=round(grand_total, 2))  # linehaul placeholder
    ws2.cell(row=5, column=12, value=0)
    ws2.cell(row=5, column=13, value=0)
    ws2.cell(row=5, column=14, value=0)
    ws2.cell(row=5, column=15, value=round(grand_total, 2))  # GRAND TOTAL — what audit reads

    ws3 = wb.create_sheet("Invoice_Lines")
    line_headers = ["#", "Invoice #", "Type", "Entry #", "Container",
                    "PO", "Description", "Qty", "Rate", "Duty", "MPF", "HMF", "Amount"]
    for col_idx, h in enumerate(line_headers, 1):
        ws3.cell(row=4, column=col_idx, value=h)
    for idx, e in enumerate(entries, 5):
        subtotal = round(
            e["duty"] + e["mpf"] + e["hmf"]
            + e["brokerage"] + e["disbursement"] + e["isf"], 2
        )
        ws3.cell(row=idx, column=1, value=idx - 4)
        ws3.cell(row=idx, column=2, value=broker_invoice_no)
        ws3.cell(row=idx, column=3, value="CUSTOMS_ENTRY")
        ws3.cell(row=idx, column=4, value=e["entry"])
        ws3.cell(row=idx, column=5, value=e["container"])
        ws3.cell(row=idx, column=6, value=e["po"])
        ws3.cell(row=idx, column=7, value=f"{e['hts']} · {e.get('notes', '')}".strip(" ·"))
        ws3.cell(row=idx, column=8, value=1)
        ws3.cell(row=idx, column=9, value=e["value"])
        ws3.cell(row=idx, column=10, value=e["duty"])
        ws3.cell(row=idx, column=11, value=e["mpf"])
        ws3.cell(row=idx, column=12, value=e["hmf"])
        ws3.cell(row=idx, column=13, value=subtotal)

    wb.save(path)
    print(f"  wrote {os.path.basename(path)}  ({len(entries)} entries · grand total ${grand_total:,.2f})")


def _build_excel(path, rows):
    """Single-sheet flat invoice list — same shape Maverick's _extract_from_excel reads."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    headers = [
        "Invoice #", "Invoice Date", "Carrier", "FB# / Load ID", "Container #",
        "BOL/MBL #", "Equipment", "Origin", "Destination",
        "Rate Type", "Base Rate", "Linehaul (USD)", "FSC %", "FSC (USD)",
        "Accessorial Detail", "Grand Total (USD)", "Currency", "Terms",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    wb.save(path)
    print(f"  wrote {os.path.basename(path)}")


def main():
    print("Generating Maverick mock test invoices...")

    # Test #1: PCD drayage — LB Pier T → Perris CA, HIGH lane, container new
    _build_pdf(
        path=os.path.join(HERE, "TEST_PCD-INV-99001.pdf"),
        header_company="PACIFIC COASTLINE DRAYAGE INC.",
        header_addr="2200 E. Anaheim Street, Long Beach, CA 90804",
        header_contact="Tel (562) 555-0190 · AP@pacificcoastlinedrayage.com · EIN 95-4827193",
        invoice_no="PCD-INV-99001",
        invoice_date="2026-05-18",
        terms="Net 30",
        fb_no="PCD25051801",
        container_no="ONEU8901234",
        bol="ONEYSZPG60270001",
        origin="Long Beach — Pier T",
        dest="NewAge Perris CA",
        equip="40HC",
        linehaul=495.00,
        fsc_pct=0.22,
        fsc_amount=108.90,
        accessorials=[("Pier Pass / Clean Truck Fee", 30.00)],
        grand_total=633.90,
        msa_ref="PCD-NA-MSA-2025-014",
    )

    # Test #2: CDS drayage — Savannah → Monee IL (will trigger different lane)
    _build_pdf(
        path=os.path.join(HERE, "TEST_CDS-INV-99001.pdf"),
        header_company="CONTINENTAL DRAYAGE SOLUTIONS",
        header_addr="980 Industrial Pkwy, Savannah, GA 31408",
        header_contact="Tel (912) 555-0181 · AP@continentaldrayage.com · EIN 58-1820439",
        invoice_no="CDS-INV-99001",
        invoice_date="2026-05-18",
        terms="Net 45",
        fb_no="CDS25051801",
        container_no="MSCU7708219",
        bol="MEDUSV80290001",
        origin="Savannah — Garden City Terminal",
        dest="NewAge Monee IL",
        equip="40HC",
        linehaul=1450.00,
        fsc_pct=0.18,
        fsc_amount=261.00,
        accessorials=[
            ("Chassis Day x 3", 90.00),
            ("Pre-pull / Storage", 175.00),  # ⚠ planted: should trigger pre-pull audit
        ],
        grand_total=1976.00,
        msa_ref="CDS-NA-MSA-2025-009",
    )

    # Test #3: Customs broker invoice (Livingston) — high-duty entry to trigger audit
    _build_customs_pdf(
        path=os.path.join(HERE, "TEST_LI-99001_customs.pdf"),
        invoice_no="LI-99001",
        invoice_date="2026-05-18",
        period="2026-05-01 to 2026-05-18",
        terms="Net 30",
        entries=[
            {
                "entry": "LI-99001-A",
                "container": "ONEU8901234",
                "po": "PO-NA-25-04481",
                "value": 184500.00,
                "hts": "9403.20.0090",
                "duty_rate": "0.0%",
                "sec301": "25%",
                "sec232": "—",
                "duty": 46125.00,
                "mpf": 538.40,
                "hmf": 235.34,
                "brokerage": 285.00,
                "disbursement": 95.00,
                "isf": 35.00,
            },
            {
                "entry": "LI-99001-B",
                "container": "MSCU7708219",
                "po": "PO-NA-25-04492",
                "value": 22400.00,
                "hts": "7321.11.6000",
                "duty_rate": "5.7%",
                "sec301": "25%",
                "sec232": "—",
                "duty": 6877.20,
                "mpf": 78.40,
                "hmf": 28.56,
                "brokerage": 175.00,
                "disbursement": 65.00,
                "isf": 35.00,
            },
        ],
        grand_total=54571.50,  # > $50k → triggers duty_math_check audit
    )

    # Test #4: Customs broker invoice in EXCEL — 10 entries (bulk customs upload)
    _build_customs_excel(
        path=os.path.join(HERE, "TEST_Customs_Bulk_10_customs.xlsx"),
        broker_invoice_no="LI-99002",
        invoice_date="2026-05-18",
        period="2026-05-04 to 2026-05-18",
        terms="Net 30",
        entries=[
            # 1 — Garage cabinets ex-China (Sec 301 stack)
            {"entry": "LI-99002-A", "container": "ONEU8901234", "po": "PO-NA-25-04481",
             "value": 184500.00, "hts": "9403.20.0090", "duty_rate": "0.0%",
             "sec301": "25%", "sec232": "—",
             "duty": 46125.00, "mpf": 538.40, "hmf": 235.34,
             "brokerage": 285.00, "disbursement": 95.00, "isf": 35.00,
             "notes": "Garage cabinet — Pro 3.0 Series"},
            # 2 — Kitchen organizers (low duty)
            {"entry": "LI-99002-B", "container": "MSCU7708219", "po": "PO-NA-25-04492",
             "value": 22400.00, "hts": "7321.11.6000", "duty_rate": "5.7%",
             "sec301": "25%", "sec232": "—",
             "duty": 6877.20, "mpf": 78.40, "hmf": 28.56,
             "brokerage": 175.00, "disbursement": 65.00, "isf": 35.00},
            # 3 — Outdoor storage cabinets (Sec 301)
            {"entry": "LI-99002-C", "container": "OOLU9912345", "po": "PO-NA-25-04503",
             "value": 145200.00, "hts": "9403.20.0090", "duty_rate": "0.0%",
             "sec301": "25%", "sec232": "—",
             "duty": 36300.00, "mpf": 485.00, "hmf": 185.21,
             "brokerage": 285.00, "disbursement": 95.00, "isf": 35.00},
            # 4 — Steel shelving (Sec 232 STEEL stack — should flag for review)
            {"entry": "LI-99002-D", "container": "HLBU4421890", "po": "PO-NA-25-04514",
             "value": 89750.00, "hts": "7308.30.5050", "duty_rate": "0.0%",
             "sec301": "25%", "sec232": "50%",
             "duty": 67312.50, "mpf": 538.40, "hmf": 114.45,
             "brokerage": 285.00, "disbursement": 95.00, "isf": 35.00,
             "notes": "STEEL — Sec 232 applies on top of Sec 301"},
            # 5 — Furniture parts
            {"entry": "LI-99002-E", "container": "TGHU5678123", "po": "PO-NA-25-04525",
             "value": 67300.00, "hts": "9403.40.9080", "duty_rate": "0.0%",
             "sec301": "25%", "sec232": "—",
             "duty": 16825.00, "mpf": 425.00, "hmf": 85.82,
             "brokerage": 225.00, "disbursement": 75.00, "isf": 35.00},
            # 6 — Cooking appliances (HS code drift flagged)
            {"entry": "LI-99002-F", "container": "COSU8765432", "po": "PO-NA-25-04536",
             "value": 98500.00, "hts": "7321.11.3000", "duty_rate": "5.7%",
             "sec301": "25%", "sec232": "—",
             "duty": 30298.50, "mpf": 538.40, "hmf": 125.62,
             "brokerage": 285.00, "disbursement": 95.00, "isf": 35.00,
             "notes": "HS classification under review — possible 7321.11.6000 drift"},
            # 7 — Mixed furniture (small)
            {"entry": "LI-99002-G", "container": "CMAU1234567", "po": "PO-NA-25-04547",
             "value": 54200.00, "hts": "9403.10.0040", "duty_rate": "0.0%",
             "sec301": "25%", "sec232": "—",
             "duty": 13550.00, "mpf": 345.00, "hmf": 69.15,
             "brokerage": 225.00, "disbursement": 75.00, "isf": 35.00},
            # 8 — Steel structures (Sec 232 again)
            {"entry": "LI-99002-H", "container": "GESU9876543", "po": "PO-NA-25-04558",
             "value": 112800.00, "hts": "7308.30.5050", "duty_rate": "0.0%",
             "sec301": "25%", "sec232": "50%",
             "duty": 84600.00, "mpf": 538.40, "hmf": 143.90,
             "brokerage": 285.00, "disbursement": 95.00, "isf": 35.00,
             "notes": "STEEL — Sec 232 applies"},
            # 9 — Small garage organizer shipment
            {"entry": "LI-99002-I", "container": "HMMU5432109", "po": "PO-NA-25-04569",
             "value": 43100.00, "hts": "9403.20.0090", "duty_rate": "0.0%",
             "sec301": "25%", "sec232": "—",
             "duty": 10775.00, "mpf": 275.00, "hmf": 54.99,
             "brokerage": 225.00, "disbursement": 75.00, "isf": 35.00},
            # 10 — Garage cabinets (overhead-priced — duty math worth auditing)
            {"entry": "LI-99002-J", "container": "ZIMU0987654", "po": "PO-NA-25-04580",
             "value": 78900.00, "hts": "9403.20.0090", "duty_rate": "0.0%",
             "sec301": "25%", "sec232": "—",
             "duty": 19725.00, "mpf": 502.85, "hmf": 100.66,
             "brokerage": 225.00, "disbursement": 75.00, "isf": 35.00},
        ],
    )

    # Test #5: Drayage Excel — two invoices in one file, including an ACS lane
    _build_excel(
        path=os.path.join(HERE, "TEST_Bulk_Invoices.xlsx"),
        rows=[
            {
                "Invoice #": "ACS-INV-99001",
                "Invoice Date": "2026-05-17",
                "Carrier": "Atlantic Container Services",
                "FB# / Load ID": "ACS25051701",
                "Container #": "HLBU4421890",
                "BOL/MBL #": "HLCUEC123456",
                "Equipment": "40HC",
                "Origin": "Charleston — Wando Welch",
                "Destination": "NewAge Monee IL",
                "Rate Type": "DRY",
                "Base Rate": 1320.00,
                "Linehaul (USD)": 1320.00,
                "FSC %": 0.18,
                "FSC (USD)": 237.60,
                "Accessorial Detail": "Chassis day x 2",
                "Grand Total (USD)": 1617.60,
                "Currency": "USD",
                "Terms": "Net 30",
            },
            {
                "Invoice #": "PCD-INV-99002",
                "Invoice Date": "2026-05-17",
                "Carrier": "Pacific Coastline Drayage",
                "FB# / Load ID": "PCD25051702",
                "Container #": "OOLU9912345",
                "BOL/MBL #": "OOLU2050001234",
                "Equipment": "40HC",
                "Origin": "Long Beach — Pier T",
                "Destination": "NewAge Perris CA",
                "Rate Type": "DRY",
                "Base Rate": 495.00,
                "Linehaul (USD)": 495.00,
                "FSC %": 0.22,
                "FSC (USD)": 108.90,
                "Accessorial Detail": "Pier Pass; Chassis day",
                "Grand Total (USD)": 663.90,
                "Currency": "USD",
                "Terms": "Net 30",
            },
        ],
    )

    print("Done.")


if __name__ == "__main__":
    main()
