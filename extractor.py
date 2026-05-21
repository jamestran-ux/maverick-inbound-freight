"""PDF invoice extractor.

Two paths:
  1. ANTHROPIC_API_KEY set -> Claude Sonnet 4 with native PDF input
  2. No key -> pdfplumber + regex deterministic parser

Both return the same JSON shape so audit logic is path-independent.
"""
import os
import re
import base64
import pdfplumber

ANTHROPIC_AVAILABLE = bool(os.environ.get("ANTHROPIC_API_KEY"))


def extract_invoice(file_path: str) -> dict:
    """Extract invoice data from a PDF or Excel file."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return _extract_from_excel(file_path)
    # default = PDF path
    if ANTHROPIC_AVAILABLE:
        try:
            return _extract_via_anthropic(file_path)
        except Exception as e:
            print(f"  [warn] Anthropic extraction failed: {e}; falling back to regex")
    return _extract_via_regex(file_path)


def _extract_from_excel(xlsx_path: str) -> dict:
    """Parse an Excel file. Supports three shapes:
       1. Flat multi-invoice list — headers on row 1, one invoice per data row
       2. Single-sheet flat invoice with side-by-side metadata cells
       3. Multi-sheet workbook (uses 'Invoices' + 'Invoice_Lines' sheets)
    Returns the same dict shape as PDF extraction. For shape (1) the primary
    record gets an 'invoices' list so callers can render N separate rows.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Try multi-sheet structure first (master workbook format)
    if "Invoices" in wb.sheetnames and "Invoice_Lines" in wb.sheetnames:
        return _extract_from_multisheet(wb)

    ws = wb.active

    # Shape (1): row 1 looks like column headers; rows 2+ are invoice rows.
    row1 = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    row1_lower = [v.lower() for v in row1]
    if "invoice #" in row1_lower and "grand total (usd)" in row1_lower:
        col = {row1_lower[i]: i + 1 for i in range(len(row1_lower))}

        def _gv(r, key):
            idx = col.get(key, 0)
            return ws.cell(row=r, column=idx).value if idx else None

        invoices_list = []
        for r in range(2, ws.max_row + 1):
            inv_no = _gv(r, "invoice #")
            if not inv_no:
                continue
            inv_no_str = str(inv_no).strip()
            if inv_no_str.lower() in ("invoice #", "invoice number"):
                continue  # belt-and-suspenders: never ingest the header row
            base = float(_gv(r, "base rate") or _gv(r, "linehaul (usd)") or 0)
            fsc_pct_raw = _gv(r, "fsc %") or 0
            try:
                fsc_pct = float(str(fsc_pct_raw).replace("%", "")) if fsc_pct_raw else 0
                if fsc_pct > 1:
                    fsc_pct = fsc_pct / 100.0
            except (TypeError, ValueError):
                fsc_pct = 0
            invoices_list.append({
                "invoice_no": inv_no_str,
                "carrier_name": str(_gv(r, "carrier") or "Unknown"),
                "invoice_date": str(_gv(r, "invoice date") or ""),
                "fb_no": str(_gv(r, "fb# / load id") or ""),
                "container_no": str(_gv(r, "container #") or ""),
                "bol": str(_gv(r, "bol/mbl #") or ""),
                "origin": str(_gv(r, "origin") or ""),
                "destination": str(_gv(r, "destination") or ""),
                "base_rate": base,
                "fsc_pct": fsc_pct,
                "fsc_amount": float(_gv(r, "fsc (usd)") or 0),
                "accessorials_total": 0.0,
                "grand_total": float(_gv(r, "grand total (usd)") or 0),
            })
        if invoices_list:
            primary = dict(invoices_list[0])
            primary["lines"] = []
            primary["confidence"] = 0.93
            primary["invoices"] = invoices_list
            return primary
        # else fall through to side-by-side metadata path
    invoice_no = None
    carrier_name = "Unknown Carrier"
    invoice_date = None
    fb_no = None
    container_no = None
    bol = None
    origin = None
    destination = None
    base_rate = 0.0
    fsc_pct = 0.0
    fsc_amount = 0.0
    grand_total = 0.0
    lines = []

    # Scan first 20 rows for header metadata
    for r in range(1, min(25, ws.max_row + 1)):
        for c in range(1, min(10, ws.max_column + 1)):
            v = ws.cell(row=r, column=c).value
            if not v: continue
            sv = str(v).strip().lower()
            nxt = ws.cell(row=r, column=c+1).value
            if "invoice #" in sv or "invoice number" in sv:
                invoice_no = str(nxt) if nxt else None
            elif "carrier" in sv and not carrier_name.startswith("Pacific"):
                carrier_name = str(nxt) if nxt else carrier_name
            elif "invoice date" in sv:
                invoice_date = str(nxt) if nxt else None

    # If sheet has a header row matching the Invoice_Lines pattern, parse lines
    header_row = None
    for r in range(1, min(15, ws.max_row + 1)):
        headers = [str(ws.cell(row=r, column=c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        if "type" in headers and "description" in headers and ("amount" in headers or "rate" in headers):
            header_row = r
            break

    if header_row:
        col_idx = {(ws.cell(row=header_row, column=c).value or "").strip().lower(): c
                   for c in range(1, ws.max_column + 1)}
        for r in range(header_row + 1, ws.max_row + 1):
            t = ws.cell(row=r, column=col_idx.get("type", 0)).value if col_idx.get("type") else None
            if not t: continue
            line_type = "SHIPMENT" if "shipment" in str(t).lower() else "ACCESSORIAL"
            desc = ws.cell(row=r, column=col_idx.get("description", 0)).value if col_idx.get("description") else None
            qty = ws.cell(row=r, column=col_idx.get("qty", 0)).value if col_idx.get("qty") else 1
            rate = ws.cell(row=r, column=col_idx.get("rate", 0)).value if col_idx.get("rate") else 0
            amount = ws.cell(row=r, column=col_idx.get("amount", 0)).value if col_idx.get("amount") else 0
            try: qty_f = float(qty or 1)
            except: qty_f = 1
            try: rate_f = float(rate or 0)
            except: rate_f = 0
            try: amount_f = float(amount or 0)
            except: amount_f = 0
            lines.append({"line_type": line_type, "description": str(desc) if desc else "",
                          "qty": qty_f, "rate": rate_f, "amount": amount_f})
            if line_type == "SHIPMENT" and not base_rate:
                base_rate = ws.cell(row=r, column=col_idx.get("linehaul", 0)).value if col_idx.get("linehaul") else rate_f
                base_rate = float(base_rate or rate_f)
                fsc_amt_cell = ws.cell(row=r, column=col_idx.get("fsc $", 0)).value if col_idx.get("fsc $") else 0
                fsc_amount = float(fsc_amt_cell or 0)
                fsc_pct_cell = ws.cell(row=r, column=col_idx.get("fsc %", 0)).value if col_idx.get("fsc %") else None
                if fsc_pct_cell:
                    s = str(fsc_pct_cell).replace("%", "").strip()
                    try: fsc_pct = float(s) / 100.0 if float(s) > 1 else float(s)
                    except: fsc_pct = 0
                if not fb_no:
                    fb_no = ws.cell(row=r, column=col_idx.get("fb#", 0)).value if col_idx.get("fb#") else None
                if not container_no:
                    container_no = ws.cell(row=r, column=col_idx.get("container #", 0)).value if col_idx.get("container #") else None

    grand_total = sum(l["amount"] for l in lines) + fsc_amount + base_rate
    accessorials_total = sum(l["amount"] for l in lines if l["line_type"] == "ACCESSORIAL")

    return {
        "invoice_no": invoice_no or os.path.basename(xlsx_path).replace(".xlsx", "").replace(".xls", ""),
        "carrier_name": carrier_name,
        "invoice_date": invoice_date,
        "fb_no": str(fb_no) if fb_no else None,
        "container_no": str(container_no) if container_no else None,
        "bol": None,
        "origin": None,
        "destination": None,
        "base_rate": base_rate,
        "fsc_pct": fsc_pct,
        "fsc_amount": fsc_amount,
        "accessorials_total": accessorials_total,
        "grand_total": grand_total,
        "lines": lines,
        "confidence": 0.88,
    }


def _extract_from_multisheet(wb) -> dict:
    """If the user uploads our master workbook (multi-sheet), extract the FIRST invoice."""
    ws_inv = wb["Invoices"]
    ws_lines = wb["Invoice_Lines"]
    inv_h = {(ws_inv.cell(row=4, column=c).value or "").strip().lower(): c for c in range(1, ws_inv.max_column + 1)}
    line_h = {(ws_lines.cell(row=4, column=c).value or "").strip().lower(): c for c in range(1, ws_lines.max_column + 1)}
    # take first invoice row
    r = 5
    inv_no = ws_inv.cell(row=r, column=inv_h.get("invoice #", 2)).value
    if not inv_no:
        return {"invoice_no": "(empty)", "carrier_name": "Unknown", "lines": [], "grand_total": 0, "confidence": 0.5}
    # build line items for that invoice — capture customs-entry fields too
    # when present (entry #, container, PO, HTS, duty, MPF, HMF) so the
    # customs upload endpoint can return per-entry rows.
    lines = []
    for lr in range(5, ws_lines.max_row + 1):
        if ws_lines.cell(row=lr, column=line_h.get("invoice #", 1)).value != inv_no:
            continue
        def _get(*keys, default=None):
            for k in keys:
                idx = line_h.get(k)
                if idx:
                    v = ws_lines.cell(row=lr, column=idx).value
                    if v is not None:
                        return v
            return default
        lines.append({
            "line_type": str(_get("type", default="SHIPMENT")),
            "description": str(_get("description", default="")),
            "qty": float(_get("qty", default=1) or 1),
            "rate": float(_get("rate", default=0) or 0),
            "amount": float(_get("amount", default=0) or 0),
            "entry": _get("entry #", "entry"),
            "container": _get("container", "container #"),
            "po": _get("po", "po #"),
            "hts": _get("hts", "hts code"),
            "duty_rate": _get("duty rate"),
            "sec301": _get("sec 301", "sec301"),
            "sec232": _get("sec 232", "sec232"),
            "duty": float(_get("duty", default=0) or 0),
            "mpf": float(_get("mpf", default=0) or 0),
            "hmf": float(_get("hmf", default=0) or 0),
            "brokerage": float(_get("brokerage", default=0) or 0),
            "notes": _get("notes", default=""),
        })
    return {
        "invoice_no": str(inv_no),
        "carrier_name": str(ws_inv.cell(row=r, column=inv_h.get("carrier", 3)).value or "Unknown"),
        "invoice_date": str(ws_inv.cell(row=r, column=inv_h.get("invoice date", 4)).value or ""),
        "fb_no": str(ws_inv.cell(row=r, column=inv_h.get("fb# / load id", 5)).value or ""),
        "container_no": str(ws_inv.cell(row=r, column=inv_h.get("container #", 6)).value or ""),
        "bol": None,
        "origin": str(ws_inv.cell(row=r, column=inv_h.get("origin", 8)).value or ""),
        "destination": str(ws_inv.cell(row=r, column=inv_h.get("destination", 9)).value or ""),
        "base_rate": float(ws_inv.cell(row=r, column=inv_h.get("linehaul (usd)", 11)).value or 0),
        "fsc_pct": 0.22,
        "fsc_amount": float(ws_inv.cell(row=r, column=inv_h.get("fsc (usd)", 13)).value or 0),
        "accessorials_total": float(ws_inv.cell(row=r, column=inv_h.get("accessorials (usd)", 14)).value or 0),
        "grand_total": float(ws_inv.cell(row=r, column=inv_h.get("grand total (usd)", 15)).value or 0),
        "lines": lines,
        "confidence": 0.95,
    }


# ----- deterministic regex path -----
def _extract_via_regex(pdf_path: str) -> dict:
    with pdfplumber.open(pdf_path) as pdf:
        all_tables = []
        for page in pdf.pages:
            all_tables.extend(page.extract_tables())
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)

    upper_text = text.upper()
    is_customs_pdf = _looks_like_customs(text, pdf_path)

    # ---- Identify line-item / entries table -----------------------------
    line_table = _find_line_table(all_tables, is_customs_pdf)

    # ---- Invoice number ------------------------------------------------
    invoice_no = _first(text, [
        # Labeled near "Invoice Number / Number / No / #"
        r"Invoice\s*(?:Number|No\.?|#)[\s:]*([A-Z0-9][A-Z0-9\-]{4,40})",
        r"Inv\s*#[\s:]*([A-Z0-9][A-Z0-9\-]{4,40})",
        # Bare patterns seen in seed + variety fixtures
        r"\b(LI-\d+(?:-[A-Z])?)\b",
        r"\b([A-Z]{3,5}-INV-\d+)\b",
        r"\b(PCD-\d+-W\d+-\d+)\b",
        r"\b([A-Z]{2,5}-\d{4}-W?\d{1,3}-\d{3,6})\b",
        r"\b([A-Z]{2,5}-CHB-\d{4}-W?\d{1,3}-\d{3,6})\b",
        r"\b([A-Z]{2,5}-DEM-\d{4}-\d{3,6})\b",
        r"\b(LI-\d+)\b",
    ])

    # ---- Carrier --------------------------------------------------------
    carrier_name = _detect_carrier(text, is_customs_pdf=is_customs_pdf)

    # ---- Dates ---------------------------------------------------------
    invoice_date = _find_date_near(text, ["Invoice Date", "Date"])
    shipment_date = _find_date_near(text, ["Shipment Date", "Pickup Date", "Load Date"])

    # Build label/value dict from 2-row column-header tables first — these are
    # the most reliable source for v4-style layouts where regex over flowing
    # text bleeds across rows.
    label_values = _extract_label_value_tables(all_tables)

    # ---- FB# / Load ID -------------------------------------------------
    fb_no = _value_from_labels(label_values, ["load id", "load no", "load number", "load #", "reference no", "reference", "fb# / load id"])
    if not fb_no:
        fb_no = _first(text, [
            r"FB#\s*/\s*Load\s*ID[\s:]*([A-Z0-9][A-Z0-9\-]{3,40})",
            r"(?:Load\s*ID|Load\s*No\.?|Load\s*#|Load\s*Number)[\s:]*([A-Z0-9][A-Z0-9\-]{3,40})",
            r"(?:Reference\s*(?:No\.?|Number|#)|\bRef\b)[\s:\-]*([A-Z0-9][A-Z0-9\-]{3,40})",
            r"(?:Pro\s*#|Job\s*#|Booking\s*#)[\s:]*([A-Z0-9][A-Z0-9\-]{3,40})",
        ])

    # ---- Container # (ISO 6346) ----------------------------------------
    container_no = _value_from_labels(label_values, ["equipment number", "container number", "container no", "container #", "container", "equipment id"])
    if container_no and not re.match(r"^[A-Z]{4}\d{7}$", container_no.upper()):
        container_no = None
    if not container_no:
        container_no = _first(text, [
            r"(?:Container|Equipment|Cont\.?|Equip\.?|Container\s*No\.?|Equipment\s*ID|Equipment\s*Number)[\s:#]*([A-Z]{4}\d{7})",
            r"\b([A-Z]{4}\d{7})\b",
        ])

    # ---- BOL / MBL -----------------------------------------------------
    bol = _value_from_labels(label_values, ["bol/mawb", "bol", "mbl", "master b/l", "b/l", "bill of lading", "b/l number"])
    if not bol:
        bol = _first(text, [
            r"(?:BOL\s*/\s*MBL|BOL/MAWB|Master\s*B/L|Master\s*BL|MBL)[\s:#]*([A-Z]{2,10}[A-Z0-9]{6,25})",
            r"(?:B/L\s*(?:Number|No\.?|#)?|BOL\s*(?:Number|No\.?|#)?)[\s:]*([A-Z]{2,10}[A-Z0-9]{6,25})",
            r"(?:Bill\s*of\s*Lading)[\s:#]*([A-Z]{2,10}[A-Z0-9]{6,25})",
        ])
    if not bol:
        # Last resort — find an ocean-BOL-shaped string different from container
        for m in re.finditer(r"\b([A-Z]{4}\d{8,15})\b", text):
            cand = m.group(1)
            if cand != container_no:
                bol = cand; break

    # ---- Origin / Destination -----------------------------------------
    origin = _clean_location(_value_from_labels(label_values, ["origin port", "origin", "from", "pickup port", "pickup"]))
    destination = _clean_location(_value_from_labels(label_values, ["destination", "delivery port", "delivery", "to", "consignee"]))
    if not origin:
        origin = _clean_location(_first(text, [
            r"(?:Origin|From|Pickup\s*Port|Pickup|From\s*Port|Port\s*of\s*Lading|POL)[\s:]*([^\n]{4,80})",
        ]))
    if not destination:
        destination = _clean_location(_first(text, [
            r"(?:Destination|To|Delivery|Delivery\s*Port|Port\s*of\s*Discharge|POD|Consignee)[\s:]*([^\n]{4,80})",
        ]))
    origin = _reject_header_word(origin)
    destination = _reject_header_word(destination)

    # ---- Multi-shipment roll-up table? (Container/Carrier/Shipment/PO/Charge layout)
    # Detect first because this signals a fundamentally different invoice shape
    # (one freight bill covering N shipments). When present, we surface the
    # per-shipment rows separately and the top-level fields reflect the FIRST
    # shipment so the existing single-shipment UI still works.
    shipments = _parse_shipment_rollup_table(all_tables)
    if shipments:
        first = shipments[0]
        # The shipment row data is the SOURCE OF TRUTH for these fields —
        # overrides any regex matches that may have caught label bleed from the
        # table header (e.g. "Destination Charge", "Container # Carrier").
        container_no = first.get("container_no") or container_no
        fb_no        = first.get("shipment_id")  or fb_no
        origin       = first.get("origin")       or origin
        destination  = first.get("destination")  or destination
        # Carrier label: broker (header) + actual carriers in rows
        carriers_in_rows = sorted({s.get("carrier") for s in shipments if s.get("carrier")})
        if carriers_in_rows and carrier_name and carrier_name not in carriers_in_rows:
            carrier_name = f"{carrier_name} ({' / '.join(carriers_in_rows[:3])})"

    # ---- Line items + totals -------------------------------------------
    base_rate = 0.0
    fsc_pct = 0.0
    fsc_amount = 0.0
    lines = []

    if line_table and is_customs_pdf:
        lines = _parse_customs_table(line_table)
    elif line_table:
        base_rate, fsc_pct, fsc_amount, lines = _parse_drayage_table(line_table)
    elif shipments:
        # Build SHIPMENT lines from the roll-up so the lines table is non-empty
        for s in shipments:
            lines.append({
                "line_type": "SHIPMENT",
                "description": f"Drayage {s.get('origin','')} → {s.get('destination','')} ({s.get('carrier','')})",
                "qty": 1.0, "rate": s.get("charge", 0), "amount": s.get("charge", 0),
                "container_no": s.get("container_no"),
                "shipment_id": s.get("shipment_id"),
                "po": s.get("po"),
                "carrier": s.get("carrier"),
            })
        base_rate = sum(s.get("charge", 0) for s in shipments)

    # If table parsing produced nothing useful, fall back to narrative scan
    # (handles invoices with no tables — "Linehaul: $X.XX" style)
    if not lines:
        base_rate, fsc_pct, fsc_amount, lines = _parse_narrative_charges(text)

    accessorials_total = sum(
        (l.get("amount") or l.get("subtotal") or 0)
        for l in lines if l.get("line_type") == "ACCESSORIAL"
    )

    grand_total = _find_grand_total(text)

    # If grand_total still 0 but line items present, sum them as a fallback
    if not grand_total and lines:
        if is_customs_pdf:
            grand_total = sum(_parse_money(l.get("subtotal") or l.get("amount") or 0) for l in lines)
        else:
            grand_total = sum(l.get("amount") or 0 for l in lines)
        grand_total = round(grand_total, 2)

    return {
        "invoice_no": invoice_no or os.path.basename(pdf_path).replace(".pdf", ""),
        "carrier_name": carrier_name,
        "invoice_date": invoice_date or shipment_date,
        "fb_no": fb_no,
        "container_no": container_no,
        "bol": bol,
        "origin": origin,
        "destination": destination,
        "base_rate": base_rate,
        "fsc_pct": fsc_pct,
        "fsc_amount": fsc_amount,
        "accessorials_total": accessorials_total,
        "grand_total": grand_total,
        "lines": lines,
        "shipments": shipments,  # populated only for multi-shipment roll-up invoices
        "po": (shipments[0].get("po") if shipments else None),
        "confidence": 0.85 if not is_customs_pdf else 0.88,
    }


# =========================================================================
# Robust helpers
# =========================================================================

# Carrier name → canonical mapping. Order matters — first hit wins. Keys are
# case-insensitive substrings; values are the display names we want to surface.
_CARRIER_HEURISTICS = [
    ("PACIFIC COASTLINE",      "Pacific Coastline Drayage"),
    ("CONTINENTAL DRAYAGE",    "Continental Drayage Solutions"),
    ("ATLANTIC CONTAINER",     "Atlantic Container Services"),
    ("PCD EXPRESS",            "PCD Express Drayage"),
    ("MAERSK LOGISTICS",       "Maersk Logistics North America"),
    ("MAERSK",                 "Maersk"),
    ("HAPAG-LLOYD",            "Hapag-Lloyd"),
    ("HAPAG LLOYD",            "Hapag-Lloyd"),
    ("CMA CGM",                "CMA CGM"),
    ("ONE ",                   "Ocean Network Express"),
    ("HMM ",                   "HMM"),
    ("OOCL",                   "OOCL"),
    ("ZIM ",                   "ZIM"),
    ("YANG MING",              "Yang Ming"),
    ("MSC ",                   "MSC"),
    ("LIVINGSTON",             "Livingston International"),
    ("ITN CUSTOMHOUSE",        "ITN Customhouse Brokers"),
    ("CUSTOMHOUSE BROKERS",    "Customhouse Brokers"),
]


def _detect_carrier(text, is_customs_pdf=False):
    upper = text.upper()
    for needle, label in _CARRIER_HEURISTICS:
        if needle in upper:
            return label
    if is_customs_pdf:
        return "Customs Broker"
    # Last-ditch: take the first non-empty line as company name if it's
    # short and looks like a header (no "$", no digits-only).
    for raw_line in (text.splitlines() or [])[:6]:
        line = raw_line.strip()
        if 3 < len(line) < 80 and "$" not in line and not line[:5].isdigit():
            # Skip generic words
            low = line.lower()
            if "invoice" in low or "bill" in low:
                continue
            # Looks plausible
            return line
    return "Unknown Carrier"


def _looks_like_customs(text, pdf_path):
    upper = text.upper()
    if "LIVINGSTON" in upper: return True
    if "BROKER INVOICE" in upper: return True
    if "CUSTOMHOUSE" in upper: return True
    if "U.S. BROKER" in upper or "US BROKER" in upper: return True
    if "ENTERED VALUE" in upper: return True
    if "HTS CODE" in upper or " HTS " in upper: return True
    if "DUTY RATE" in upper: return True
    if "ENTRY SUMMARY" in upper or "ENTRY NO" in upper: return True
    if "IMPORTER OF RECORD" in upper: return True
    if "customs" in os.path.basename(pdf_path).lower(): return True
    return False


# Month names → 2-digit
_MONTHS = {m: i for i, m in enumerate(
    ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"], start=1
)}


def _find_date_near(text, labels):
    """Find the first date appearing after any of these labels. Normalize to YYYY-MM-DD."""
    for label in labels:
        # Allow up to ~30 chars / 2 lines between label and date
        m = re.search(
            rf"{re.escape(label)}[\s:]*(?:\n\s*)?([A-Za-z0-9\-/, ]{{6,30}})",
            text, re.IGNORECASE,
        )
        if not m: continue
        candidate = m.group(1).strip()
        norm = _normalize_date(candidate)
        if norm:
            return norm
    # Fallback: any obvious date anywhere in text
    return _normalize_date(text)


def _normalize_date(s):
    """Accept various human-readable date forms, return YYYY-MM-DD (or None)."""
    if not s: return None
    s = s.strip()

    # YYYY-MM-DD
    m = re.search(r"\b(\d{4})-(\d{1,2})-(\d{1,2})\b", s)
    if m:
        y, mo, d = m.group(1), m.group(2).zfill(2), m.group(3).zfill(2)
        return f"{y}-{mo}-{d}"

    # MM/DD/YYYY or M/D/YYYY (also accepts 2-digit year)
    m = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\b", s)
    if m:
        mo, d, y = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        if len(y) == 2:
            y = "20" + y if int(y) < 70 else "19" + y
        return f"{y}-{mo}-{d}"

    # DD-Mon-YYYY (e.g. "19-May-2026")
    m = re.search(r"\b(\d{1,2})-([A-Za-z]{3,9})-(\d{2,4})\b", s)
    if m:
        d, mon, y = m.group(1).zfill(2), m.group(2)[:3].lower(), m.group(3)
        if mon in _MONTHS:
            if len(y) == 2:
                y = "20" + y
            return f"{y}-{_MONTHS[mon]:02d}-{d}"

    # "May 19, 2026" / "May 19 2026"
    m = re.search(r"\b([A-Za-z]{3,9})\s+(\d{1,2}),?\s+(\d{4})\b", s)
    if m:
        mon, d, y = m.group(1)[:3].lower(), m.group(2).zfill(2), m.group(3)
        if mon in _MONTHS:
            return f"{y}-{_MONTHS[mon]:02d}-{d}"

    return None


def _first(text, patterns):
    """Try each pattern; return the first matching group (stripped)."""
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
        if m:
            val = m.group(1).strip()
            # Reject obvious labels accidentally picked up
            if val and not _looks_like_label(val):
                return val
    return None


def _looks_like_label(s):
    """Reject 'INVOICE', 'DATE', etc. — these are field labels masquerading as values."""
    if not s: return True
    bare = s.strip().rstrip(":").lower()
    return bare in {"invoice", "date", "number", "no", "load", "container", "bol",
                    "mbl", "ref", "reference", "origin", "destination", "from",
                    "to", "pickup", "delivery", "total", "amount", "due"}


def _clean_location(s):
    if not s: return s
    return s.strip().rstrip(".· →").strip()


def _find_grand_total(text):
    """Best-effort grand-total finder. Tries labeled patterns, picks the
    largest reasonable USD value if multiple match."""
    candidates = []
    for label in [
        r"GRAND\s*TOTAL",
        r"TOTAL\s*AMOUNT\s*DUE",
        r"AMOUNT\s*DUE",
        r"INVOICE\s*TOTAL",
        r"BALANCE\s*DUE",
        r"TOTAL\s*DUE",
        r"\bTOTAL\b",
    ]:
        for m in re.finditer(
            rf"{label}[\s:]*\$?\s*([\d,]+\.\d{{2}})",
            text, re.IGNORECASE,
        ):
            try:
                candidates.append(float(m.group(1).replace(",", "")))
            except ValueError:
                pass
    if not candidates:
        return 0.0
    # Pick the largest match — usually that's the grand total, not a per-line
    return max(candidates)


# ----- Table parsing -----

def _find_line_table(tables, is_customs):
    """Return the most-likely line-items / entries table.

    For drayage: header starts with #/Line/Item AND has at least 4 columns.
    For customs: header includes Entry/Container/Duty-ish words.
    """
    def header_score(row, customs):
        cells = [str(c or "").strip().lower() for c in row]
        score = 0
        if customs:
            for needle in ("entry", "container", "duty", "hts", "tariff", "subtotal"):
                if any(needle in c for c in cells):
                    score += 1
        else:
            if cells and cells[0] in {"#", "line", "item"}:
                score += 2
            for needle in ("description", "qty", "rate", "amount", "charge"):
                if any(needle in c for c in cells):
                    score += 1
        return score

    best = None
    best_score = 0
    for t in tables:
        if not t or len(t) < 2:
            continue
        score = header_score(t[0], is_customs)
        if score > best_score:
            best = t
            best_score = score
    return best if best_score >= 2 else None


def _parse_drayage_table(table):
    """Parse a drayage line-items table, returning (base_rate, fsc_pct, fsc_amt, lines)."""
    header = [str(c or "").strip().lower() for c in table[0]]
    # Locate columns by header name (best-effort)
    def col(*needles):
        for i, c in enumerate(header):
            if any(n in c for n in needles):
                return i
        return None

    c_desc = col("description", "charge")
    c_qty  = col("qty", "units")
    c_rate = col("rate")
    c_amt  = col("amount", "subtotal")

    base_rate = 0.0
    fsc_pct   = 0.0
    fsc_amt   = 0.0
    lines     = []

    for row in table[1:]:
        cells = [(c or "").replace("\n", " ").strip() for c in row]
        if not any(cells): continue
        if c_desc is None or c_desc >= len(cells): continue
        desc = cells[c_desc]
        if not desc: continue

        desc_low = desc.lower()
        is_shipment = any(k in desc_low for k in ("shipment", "drayage", "linehaul", "line haul", "port to dc"))
        is_fsc = "fuel" in desc_low or "fsc" in desc_low
        # FSC is its own bucket so accessorials_total doesn't double-count it
        line_type = "FSC" if is_fsc else ("SHIPMENT" if is_shipment else "ACCESSORIAL")

        rate = _parse_money(cells[c_rate]) if c_rate is not None and c_rate < len(cells) else 0
        amt  = _parse_money(cells[c_amt])  if c_amt  is not None and c_amt  < len(cells) else 0
        try:
            qty = float(cells[c_qty]) if c_qty is not None and c_qty < len(cells) and cells[c_qty] not in ("", "—") else 1
        except ValueError:
            qty = 1

        lines.append({
            "line_type": line_type, "description": desc,
            "qty": qty, "rate": rate, "amount": amt,
        })

        if is_shipment and not is_fsc:
            base_rate = base_rate or rate or amt
        if is_fsc:
            fsc_amt = fsc_amt or amt
            # Pull "22%" or "22.0%" out of the description
            pm = re.search(r"(\d{1,3}(?:\.\d+)?)\s*%", desc)
            if pm:
                try: fsc_pct = float(pm.group(1)) / 100.0
                except ValueError: pass

    return base_rate, fsc_pct, fsc_amt, lines


def _parse_customs_table(table):
    """Parse a customs/broker entries table. Locate each field by header name
    so column reordering doesn't break the parse."""
    header = [str(c or "").strip().lower() for c in table[0]]

    def col(*needles):
        for i, c in enumerate(header):
            if any(n in c for n in needles):
                return i
        return None

    c_entry     = col("entry")
    c_container = col("container", "equipment")
    c_po        = col("po", "importer po", "purchase order")
    c_value     = col("entered value", "value")
    c_hts       = col("hts", "tariff")
    c_rate      = col("duty rate")
    c_sec301    = col("sec 301", "section 301", "301")
    c_sec232    = col("sec 232", "section 232", "232")
    c_duty      = col("duty")  # last-resort — overlaps with rate; resolve below
    c_mpf       = col("mpf")
    c_hmf       = col("hmf")
    c_brok      = col("brokerage", "broker fee")
    c_subtotal  = col("subtotal", "total")
    # Disambiguate "Duty" vs "Duty Rate" (we want the AMOUNT column, not the rate)
    if c_duty is not None and c_duty == c_rate:
        for i, c in enumerate(header):
            if c == "duty":
                c_duty = i; break

    lines = []
    for row in table[1:]:
        cells = [(c or "").replace("\n", " ").strip() for c in row]
        if not any(cells): continue
        entry = cells[c_entry] if c_entry is not None and c_entry < len(cells) else ""
        if not entry: continue
        def cell(idx):
            return cells[idx] if idx is not None and idx < len(cells) else ""
        lines.append({
            "line_type": "CUSTOMS_ENTRY",
            "entry":      entry,
            "container":  cell(c_container),
            "po":         cell(c_po),
            "value":      _parse_money(cell(c_value)),
            "hts":        cell(c_hts),
            "duty_rate":  cell(c_rate),
            "sec301":     cell(c_sec301) or "—",
            "sec232":     cell(c_sec232) or "—",
            "duty":       _parse_money(cell(c_duty)),
            "mpf":        _parse_money(cell(c_mpf)),
            "hmf":        _parse_money(cell(c_hmf)),
            "brokerage":  _parse_money(cell(c_brok)),
            "amount":     _parse_money(cell(c_subtotal)),
            "subtotal":   _parse_money(cell(c_subtotal)),
            "description": cell(c_hts) + " · " + entry,
            "qty": 1.0,
            "rate": _parse_money(cell(c_value)),
        })
    return lines


def _parse_shipment_rollup_table(tables):
    """Detect the "multi-shipment roll-up" layout where one invoice covers
    several shipments and each row is its own shipment with container,
    carrier, shipment id, PO, lane, and charge.

    Header signature (any subset must include Container AND (Shipment OR PO)
    AND a money/charge column):
        Container # | Carrier | Shipment ID | PO # | Origin | Destination | Charge
    Returns a list of shipment dicts, or [] if no matching table is found.
    """
    for t in tables:
        if not t or len(t) < 2: continue
        header = [str(c or "").strip().lower() for c in t[0]]
        if not any("container" in c for c in header): continue
        has_shipment = any("shipment" in c for c in header)
        has_po       = any(("po" in c) or ("purchase order" in c) for c in header)
        has_money    = any(c in {"charge", "amount", "subtotal", "total", "rate"} for c in header)
        if not (has_shipment or has_po): continue
        if not has_money: continue

        def col(*needles):
            for i, c in enumerate(header):
                if any(n in c for n in needles): return i
            return None

        c_cont   = col("container")
        c_carr   = col("carrier")
        c_ship   = col("shipment")
        c_po     = col("po", "purchase order")
        c_orig   = col("origin", "from", "pol", "pickup")
        c_dest   = col("destination", "delivery", "pod", "to ", "consignee")
        c_money  = col("charge", "amount", "subtotal", "total", "rate")

        shipments = []
        for row in t[1:]:
            cells = [(c or "").replace("\n", " ").strip() for c in row]
            if not any(cells): continue
            def cell(idx):
                return cells[idx] if idx is not None and idx < len(cells) else ""
            container = cell(c_cont)
            # Skip total-style rows ("Subtotal", "Total Due", etc.)
            if not container or container.lower() in ("subtotal", "total", "total due", "gst"): continue
            if not re.match(r"^[A-Z]{4}\d{7}$", container.upper()):
                # Container column doesn't look like ISO 6346 — bail on this row
                continue
            shipments.append({
                "container_no": container,
                "carrier":     cell(c_carr) or None,
                "shipment_id": cell(c_ship) or None,
                "po":          cell(c_po) or None,
                "origin":      cell(c_orig) or None,
                "destination": cell(c_dest) or None,
                "charge":      _parse_money(cell(c_money)),
            })
        if shipments:
            return shipments
    return []


def _parse_narrative_charges(text):
    """Fallback for invoices with no tables: scrape charge lines from prose.
    Recognizes bullets •, -, *, en/em-dash, AND (cid:127) (ReportLab bullet
    glyph that pdfplumber can't decode). FSC lines are tagged "FSC" so they
    don't double-count as accessorials.
    """
    base_rate = 0.0
    fsc_pct   = 0.0
    fsc_amt   = 0.0
    lines     = []

    # Strip ReportLab's undecoded cid markers so the bullet doesn't break the regex
    cleaned = re.sub(r"\(cid:\d+\)", "", text)

    pattern = re.compile(
        r"(?:^|\n)[•\-\*‒‐–—]?\s*([A-Z][A-Za-z][^:\n]{2,80}?)[:\s]+\$\s*([\d,]+\.\d{2})",
        re.MULTILINE,
    )
    for m in pattern.finditer(cleaned):
        desc = m.group(1).strip()
        amt  = _parse_money(m.group(2))
        if re.search(r"\b(grand\s*total|total\s*amount|amount\s*due|balance\s*due|invoice\s*total|total\s*due|\btotal\b)\b",
                     desc, re.IGNORECASE):
            continue
        is_shipment = bool(re.search(r"linehaul|line haul|drayage|shipment", desc, re.IGNORECASE))
        is_fsc = bool(re.search(r"fuel|fsc", desc, re.IGNORECASE))
        line_type = "FSC" if is_fsc else ("SHIPMENT" if is_shipment else "ACCESSORIAL")
        lines.append({
            "line_type": line_type, "description": desc,
            "qty": 1, "rate": amt, "amount": amt,
        })
        if is_shipment and not is_fsc:
            base_rate = base_rate or amt
        if is_fsc:
            fsc_amt = fsc_amt or amt
            pm = re.search(r"(\d{1,3}(?:\.\d+)?)\s*%", desc)
            if pm:
                try: fsc_pct = float(pm.group(1)) / 100.0
                except ValueError: pass

    return base_rate, fsc_pct, fsc_amt, lines


# ---- Label/value table helpers ------------------------------------------
def _extract_label_value_tables(tables):
    """Find 2-row tables where row 0 is labels and row 1 is values. Return a
    label-lowercase → value dict (last-write wins)."""
    out = {}
    for t in tables:
        if not t or len(t) < 2: continue
        # Only accept 2-row tables (label + value) — line tables have >2 rows
        if len(t) != 2: continue
        header = [str(c or "").strip() for c in t[0]]
        values = [str(c or "").strip() for c in t[1]]
        if len(header) != len(values): continue
        # Require at least one cell to look like a label (not a money/date value)
        label_like = sum(1 for h in header if h and not re.search(r"\d", h))
        if label_like < max(1, len(header) // 2): continue
        for label, value in zip(header, values):
            if not label or not value: continue
            out[label.lower().strip().rstrip(":")] = value.strip()
    return out


def _value_from_labels(d, labels):
    """Lookup the first label key (lowercase) that exists in d."""
    for lab in labels:
        v = d.get(lab.lower())
        if v: return v
    return None


def _reject_header_word(s):
    """If the captured string is just all-uppercase header-style words (e.g.
    'PORT DESTINATION'), reject it as a label collision."""
    if not s: return s
    # All-caps with no digits and no lowercase → looks like a header word run
    if re.match(r"^[A-Z][A-Z\s/&\.\-]+$", s.strip()) and not re.search(r"\d", s):
        # Reject if it's mostly recognized label tokens
        tokens = re.split(r"\s+", s.strip())
        labels = {"PORT", "DESTINATION", "ORIGIN", "BOL", "MBL", "MAWB", "EQUIPMENT",
                  "NUMBER", "INVOICE", "DATE", "LOAD", "ID", "ITEM", "CHARGE", "UNITS",
                  "RATE", "SUBTOTAL", "AMOUNT", "DUE"}
        if all(t in labels for t in tokens if t):
            return None
    return s


def _re1(pattern, text):
    m = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
    return m.group(1).strip() if m else None


def _parse_money(s):
    if s is None or s == "—":
        return 0.0
    s = str(s).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


# ----- AI path -----
def _extract_via_anthropic(pdf_path):
    import anthropic
    client = anthropic.Anthropic()
    with open(pdf_path, "rb") as f:
        pdf_b64 = base64.b64encode(f.read()).decode("utf-8")
    schema = {
        "type": "object",
        "properties": {
            "invoice_no": {"type": "string", "description": "Invoice number / freight bill ID at top of doc"},
            "carrier_name": {"type": "string", "description": "Issuing carrier or broker name"},
            "invoice_date": {"type": "string", "description": "Invoice date in YYYY-MM-DD format if possible"},
            "fb_no": {"type": "string", "description": "FB# / Load ID / Shipment reference for the (first) shipment"},
            "container_no": {"type": "string", "description": "ISO 6346 container number (4 letters + 7 digits) for the (first) shipment"},
            "bol": {"type": "string", "description": "Bill of Lading / Master B/L / MBL number"},
            "po": {"type": "string", "description": "Purchase Order number for the (first) shipment"},
            "origin": {"type": "string"},
            "destination": {"type": "string"},
            "base_rate": {"type": "number"},
            "fsc_pct": {"type": "number"},
            "fsc_amount": {"type": "number"},
            "accessorials_total": {"type": "number"},
            "grand_total": {"type": "number"},
            "lines": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "line_type": {"type": "string", "enum": ["SHIPMENT", "ACCESSORIAL", "FSC"]},
                        "description": {"type": "string"},
                        "qty": {"type": "number"},
                        "rate": {"type": "number"},
                        "amount": {"type": "number"},
                    },
                    "required": ["line_type", "description", "qty", "rate", "amount"],
                },
            },
            "shipments": {
                "type": "array",
                "description": "When the invoice is a freight bill covering MULTIPLE shipments (each row of an embedded shipment table is its own shipment), populate one entry per shipment. Otherwise leave empty.",
                "items": {
                    "type": "object",
                    "properties": {
                        "container_no": {"type": "string", "description": "ISO 6346 container # for this shipment"},
                        "shipment_id":  {"type": "string", "description": "Shipment ID / SHIP-XXX / FB# for this shipment"},
                        "po":           {"type": "string", "description": "Purchase Order # for this shipment"},
                        "carrier":      {"type": "string", "description": "Underlying carrier (CN Rail, CPKC, etc.) — leave empty if same as invoice carrier"},
                        "origin":       {"type": "string"},
                        "destination":  {"type": "string"},
                        "charge":       {"type": "number"},
                    },
                    "required": ["container_no", "charge"],
                },
            },
        },
        "required": ["invoice_no", "carrier_name", "lines", "grand_total"],
    }
    response = client.messages.create(
        model="claude-sonnet-4-5-20250929",
        max_tokens=4000,
        tools=[{"name": "submit_invoice", "description": "Submit the parsed invoice data",
                "input_schema": schema}],
        tool_choice={"type": "tool", "name": "submit_invoice"},
        messages=[{
            "role": "user",
            "content": [
                {"type": "document",
                 "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                {"type": "text", "text":
                 "Extract this drayage freight invoice into the structured schema. "
                 "Rules:\n"
                 "1. Container numbers are ISO 6346 format: 4 letters + 7 digits (e.g. MSCU5990314, HLCU7936445).\n"
                 "2. If the invoice contains a SHIPMENT TABLE with a row per shipment "
                 "(headers like 'Container # | Shipment ID | PO # | Origin | Destination | Charge'), "
                 "populate the `shipments` array with one entry PER ROW. Each shipment has its own "
                 "container, shipment_id, PO, origin, destination, charge. Also populate top-level "
                 "container_no/fb_no/po/origin/destination/base_rate from the FIRST shipment.\n"
                 "3. For single-shipment invoices, leave `shipments` empty and populate only top-level fields.\n"
                 "4. fsc_pct is a fraction 0-1 (so 22% = 0.22). Tag fuel surcharge lines as line_type 'FSC' "
                 "(not 'ACCESSORIAL') so they don't double-count.\n"
                 "5. Carrier names: prefer the issuing carrier (top of doc), not the underlying ocean line."},
            ],
        }],
    )
    for block in response.content:
        if block.type == "tool_use" and block.name == "submit_invoice":
            data = dict(block.input)
            data["confidence"] = 0.95
            data.setdefault("shipments", [])
            # Safety net: when AI returned shipments[] but left the top-level
            # container/fb/po blank, copy from the first shipment so the
            # invoice row inserted into D.invoices isn't empty.
            ships = data.get("shipments") or []
            if ships:
                first = ships[0] or {}
                if not data.get("container_no"): data["container_no"] = first.get("container_no")
                if not data.get("fb_no"):        data["fb_no"]        = first.get("shipment_id")
                if not data.get("po"):           data["po"]           = first.get("po")
                if not data.get("origin"):       data["origin"]       = first.get("origin")
                if not data.get("destination"):  data["destination"]  = first.get("destination")
                if not data.get("base_rate") and first.get("charge"): data["base_rate"] = first.get("charge")
            return data
    raise RuntimeError("Anthropic returned no tool_use block")


if __name__ == "__main__":
    import sys, json
    if len(sys.argv) < 2:
        print("Usage: python extractor.py <pdf_path>")
        sys.exit(1)
    result = extract_invoice(sys.argv[1])
    print(json.dumps(result, indent=2, default=str))
