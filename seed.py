"""Seed the database from Mock_Drayage_Invoice_Data.xlsx.

Idempotent — uses INSERT OR REPLACE keyed on natural primary keys so running
twice doesn't duplicate.
"""
import os
import re
import openpyxl
from datetime import datetime
from db import get_conn

_HERE = os.path.dirname(os.path.abspath(__file__))
# Look in the repo first (for cloud deploys), then in the parent folder (local dev)
_LOCAL = os.path.join(_HERE, "Mock_Drayage_Invoice_Data.xlsx")
_PARENT = os.path.join(os.path.dirname(_HERE), "Mock_Drayage_Invoice_Data.xlsx")
EXCEL_PATH = _LOCAL if os.path.exists(_LOCAL) else _PARENT


def _val(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return v


def _pct(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v) if v <= 1 else float(v) / 100.0
    s = str(v).replace("%", "").strip()
    try:
        f = float(s)
        return f / 100.0 if f > 1 else f
    except ValueError:
        return 0.0


def _percent_str_to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("%", "").strip()
    if not s or s == "—":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _header_map(ws, header_row):
    return {(c.value or "").strip().lower(): i for i, c in enumerate(ws[header_row])}


def _rows_after(ws, header_row):
    return ws.iter_rows(min_row=header_row + 1, values_only=True)


def seed_all(excel_path=EXCEL_PATH):
    print(f"Seeding from {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    conn = get_conn()
    try:
        seed_carriers(conn, wb)
        seed_carrier_scorecard(conn, wb)
        seed_rate_card(conn, wb)
        seed_accessorial_rates(conn, wb)
        seed_terminal_appointments(conn, wb)
        seed_carrier_capacity(conn, wb)
        seed_per_diem_ladder(conn, wb)
        seed_containers(conn, wb)
        seed_pos(conn, wb)
        seed_customs_invoices(conn, wb)
        seed_transfers(conn, wb)
        seed_p4_transfer_needs(conn)
        seed_users(conn)
        seed_invoices_and_lines(conn, wb)
        seed_loads_from_invoices_and_containers(conn)
        seed_gl_accruals(conn)
        conn.commit()
    finally:
        conn.close()
    print("Seed complete.")


VALID_CARRIER_TYPES = {"Third-Party Dray", "Third-Party Dray (Premium)",
                       "Ocean Carrier Dray", "Rail Carrier Dray"}


def seed_carriers(conn, wb):
    ws = wb["Carrier_Scorecard"]
    h = _header_map(ws, 4)
    seen = set()
    for row in _rows_after(ws, 4):
        if not row or not row[h["carrier"]]:
            continue
        carrier_type = row[h["carrier type"]]
        if carrier_type not in VALID_CARRIER_TYPES:
            # skip legend rows
            continue
        name = row[h["carrier"]]
        if name in seen:
            continue
        seen.add(name)
        slug = re.sub(r"[^a-zA-Z]", "", name).lower()
        conn.execute(
            "INSERT OR REPLACE INTO carriers (name, carrier_type, contact_email) VALUES (?, ?, ?)",
            (name, carrier_type, f"dispatch@{slug}.com"),
        )


def seed_carrier_scorecard(conn, wb):
    ws = wb["Carrier_Scorecard"]
    h = _header_map(ws, 4)
    for row in _rows_after(ws, 4):
        if not row or not row[h["carrier"]]:
            continue
        if row[h["carrier type"]] not in VALID_CARRIER_TYPES:
            continue
        conn.execute(
            """INSERT OR REPLACE INTO carrier_scorecard (carrier_name, carrier_type, on_time_pickup,
                on_time_delivery, invoice_accuracy, accessorial_pct, dispute_win_rate,
                composite_score, trailing_90d_loads, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                row[h["carrier"]],
                row[h["carrier type"]],
                _percent_str_to_float(row[h["on-time pickup %"]]),
                _percent_str_to_float(row[h["on-time delivery %"]]),
                _percent_str_to_float(row[h["invoice accuracy %"]]),
                _percent_str_to_float(row[h["accessorial spend % of base"]]),
                _percent_str_to_float(row[h["dispute win rate %"]]),
                row[h["composite score"]],
                row[h["trailing 90d loads"]],
                row[h["notes"]],
            ),
        )


def seed_rate_card(conn, wb):
    ws = wb["Rate_Card"]
    h = _header_map(ws, 4)
    for row in _rows_after(ws, 4):
        if not row or not row[h["lane id"]]:
            continue
        lane = str(row[h["lane id"]])
        if not lane.startswith("RC-"):
            continue
        try:
            conn.execute(
                """INSERT OR REPLACE INTO rate_card (lane_id, carrier_name, carrier_type,
                    origin_terminal, destination_dc, lane_criticality, equipment, base_rate,
                    fsc_pct, tier, effective_from, effective_to, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    row[h["lane id"]], row[h["carrier"]], row[h["carrier type"]],
                    row[h["origin port/terminal"]], row[h["destination dc"]],
                    row[h["lane criticality"]], row[h["equipment"]],
                    float(row[h["base rate (usd)"]]),
                    _pct(row[h["fsc %"]]),
                    row[h["tier"]],
                    _val(row[h["effective from"]]),
                    _val(row[h["effective to"]]),
                    row[h["notes"]],
                ),
            )
        except (TypeError, ValueError, KeyError):
            continue


def seed_accessorial_rates(conn, wb):
    ws = wb["Rate_Card"]
    acc_start = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "ACCESSORIAL RATE CARD":
            acc_start = r + 2
            break
    if acc_start is None:
        return
    h = _header_map(ws, acc_start)
    for row in ws.iter_rows(min_row=acc_start + 1, values_only=True):
        if not row or not row[h["code"]]:
            continue
        rate_val = row[h["rate (usd)"]]
        try:
            rate = float(rate_val) if isinstance(rate_val, (int, float)) else None
        except (TypeError, ValueError):
            rate = None
        conn.execute(
            """INSERT OR REPLACE INTO accessorial_rates (code, description, unit, rate, free_allowance, notes)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (row[h["code"]], row[h["description"]], row[h["unit"]], rate,
             row[h["free allowance"]], row[h["notes"]]),
        )


def seed_terminal_appointments(conn, wb):
    ws = wb["Terminal_Appointments"]
    h = _header_map(ws, 4)
    for row in _rows_after(ws, 4):
        if not row or not row[h["terminal"]]:
            continue
        conn.execute(
            """INSERT INTO terminal_appointments (terminal, equipment, next_available_date,
                window, open_slots, avg_wait_min, system, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (row[h["terminal"]], row[h["equipment"]],
             _val(row[h["next available date"]]), row[h["window"]],
             row[h["open slots"]], row[h["avg wait (min)"]],
             row[h["system"]], row[h["notes"]]),
        )


def seed_carrier_capacity(conn, wb):
    ws = wb["Carrier_Capacity"]
    h = _header_map(ws, 4)
    for row in _rows_after(ws, 4):
        if not row or not row[h["carrier"]]:
            continue
        conn.execute(
            """INSERT INTO carrier_capacity (carrier_name, lane_group, equipment,
                weekly_capacity, this_week_committed, available, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (row[h["carrier"]], row[h["lane group"]], row[h["equipment"]],
             row[h["weekly capacity"]], row[h["this week committed"]],
             row[h["available"]], row[h["notes"]]),
        )


def seed_per_diem_ladder(conn, wb):
    ws = wb["Per_Diem_Ladder"]
    h = _header_map(ws, 4)
    for row in _rows_after(ws, 4):
        if not row or not row[h["steamship line"]]:
            continue
        conn.execute(
            """INSERT INTO per_diem_ladder (steamship_line, equipment, free_time_port,
                free_time_consignee, demurrage_d1_3, demurrage_d4_7, demurrage_d8_plus,
                detention_d1_3, detention_d4_7, detention_d8_plus, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (row[h["steamship line"]], row[h["equipment"]],
             row[h["free time at port (days)"]], row[h["free time at consignee (days)"]],
             row[h["demurrage days 1–3 ($/day)"]], row[h["demurrage days 4–7 ($/day)"]],
             row[h["demurrage days 8+ ($/day)"]],
             row[h["detention days 1–3 ($/day)"]], row[h["detention days 4–7 ($/day)"]],
             row[h["detention days 8+ ($/day)"]],
             row[h["notes"]]),
        )


def seed_containers(conn, wb):
    ws = wb["Containers"]
    h = _header_map(ws, 4)
    for row in _rows_after(ws, 4):
        if not row or not row[h["container #"]]:
            continue
        conn.execute(
            """INSERT OR REPLACE INTO containers (number, steamship_line, vessel, equipment,
                origin_port, us_port, discharge_date, customs_status, ssl_released,
                lfd, pickup_date, free_time_days, stage, container_status, days_at_location,
                demurrage_risk, status, linked_po, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                row[h["container #"]], row[h["steamship line"]], row[h["vessel"]],
                row[h["equipment"]], row[h["origin port"]], row[h["us port"]],
                _val(row[h["discharge date"]]), row[h["customs status"]],
                row[h["ssl released"]],
                _val(row[h["lfd"]]), _val(row[h["pickup date"]]),
                row[h["free time (days)"]],
                row[h["stage"]], row[h["container status"]],
                row[h["days at location"]] if isinstance(row[h["days at location"]], (int, float)) else None,
                row[h["demurrage risk"]], row[h["status"]],
                row[h["linked po"]], row[h["notes"]],
            ),
        )


def seed_pos(conn, wb):
    ws = wb["POs"]
    h = _header_map(ws, 3)
    for row in _rows_after(ws, 3):
        if not row or not row[h["po #"]]:
            continue
        conn.execute(
            """INSERT OR REPLACE INTO purchase_orders (po_no, supplier, origin, sku_family,
                units, unit_cost, value, hts_code, container_no, issue_date, eta)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (row[h["po #"]], row[h["supplier"]], row[h["origin"]], row[h["sku family"]],
             row[h["units"]], row[h["unit cost (usd)"]], row[h["value (usd)"]],
             row[h["hts code"]], row[h["container #"]],
             _val(row[h["issue date"]]), _val(row[h["eta"]])),
        )


def seed_customs_invoices(conn, wb):
    ws = wb["Customs_Invoice"]
    hdr_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Entry #":
            hdr_row = r
            break
    if hdr_row is None:
        return
    h = _header_map(ws, hdr_row)
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if not row or not row[h["entry #"]]:
            continue
        try:
            conn.execute(
                """INSERT OR REPLACE INTO customs_invoices (entry_no, broker_name, invoice_date,
                    container_no, po_no, entered_value, hts_code, duty_rate, sec301_pct,
                    sec232_pct, duty, mpf, hmf, brokerage, disbursement, isf, subtotal, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (row[h["entry #"]], "Livingston International", "2026-05-12",
                 row[h["container #"]], row[h["po #"]], row[h["entered value (usd)"]],
                 row[h["hts code"]], str(row[h["duty rate"]]) if row[h["duty rate"]] is not None else None,
                 str(row[h["section 301"]]) if row[h["section 301"]] is not None else None,
                 str(row[h["section 232"]]) if row[h["section 232"]] is not None else None,
                 row[h["duty (usd)"]], row[h["mpf (usd)"]], row[h["hmf (usd)"]],
                 row[h["brokerage (usd)"]], row[h["disbursement (usd)"]],
                 row[h["isf (usd)"]], row[h["subtotal (usd)"]],
                 row[h["notes"]]),
            )
        except (TypeError, KeyError):
            continue


def seed_transfers(conn, wb):
    ws = wb["Transfers"]
    h = _header_map(ws, 4)
    for row in _rows_after(ws, 4):
        if not row or not row[h["transfer id"]]:
            continue
        conn.execute(
            """INSERT OR REPLACE INTO transfer_requests (transfer_id, from_dc, to_dc, mode,
                equipment, equipment_ref, sku_family, reason, need_by, est_cost, actual_cost,
                status, costco_po_ref, requested_by, approved_by, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (row[h["transfer id"]], row[h["from dc"]], row[h["to dc"]], row[h["mode"]],
             row[h["equipment"]], row[h["container #"]], row[h["sku family"]],
             row[h["reason"]],
             _val(row[h["need-by"]]),
             row[h["est. cost (usd)"]], row[h["actual cost (usd)"]],
             row[h["status"]], row[h["costco po linked?"]],
             row[h["requested by"]], row[h["approved by"]], row[h["notes"]]),
        )


def seed_p4_transfer_needs(conn):
    needs = [
        ("P4-REQ-77419", "Perris CA", "Monee IL", "Garage Cabinet — Pro 3.0 Series",
         1200, "2026-05-23", "Costco PO #C-77419", "Costco Naperville POG drop", "Pending"),
        ("P4-REQ-77420", "Monee IL", "Perris CA", "Outdoor Kitchen — SS Cabinet",
         340, "2026-05-21", None, "Home Depot SoCal stockout", "Pending"),
        ("P4-REQ-77421", "Perris CA", "Monee IL", "BBQ Gas Grill — Platinum Series",
         180, "2026-05-26", None, "Recall quarantine", "Pending"),
        ("P4-REQ-77422", "LB Pier T", "Monee IL", "Garage Cabinet — Pro 3.0 Series",
         1400, "2026-05-24", None, "Direct import — bypass Perris", "Pending"),
        ("P4-REQ-77423", "Perris CA", "Monee IL", "Home Bar Cabinet — Maple Series",
         420, "2026-05-28", None, "Network rebalance — forecasted demand", "Pending"),
    ]
    for n in needs:
        conn.execute(
            """INSERT OR REPLACE INTO p4_transfer_needs (p4_request_id, from_dc, to_dc,
                sku_family, units, need_by, costco_po_ref, reason, status)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            n,
        )


def seed_users(conn):
    users = [
        ("James", "Tran", "james.tran@newageproducts.com", "Inbound Freight Manager", "Today"),
        ("Alexander", "Curlat-Rozenberg", "alexander.curlat-rozenberg@newageproducts.com",
         "Director of Logistics", "Today"),
        ("Alec", "Swindeman", "alec.swindeman@newageproducts.com",
         "Sr Manager, Supply Chain", "Yesterday"),
        ("Rahul", "Sharma", "rahul.sharma@newageproducts.com",
         "Director, Demand Planning", "2 days ago"),
        ("Akshay", "Kapasi", "akshay.kapasi@newageproducts.com",
         "Global Sourcing Lead", "1 week ago"),
        ("Demo", "Finance", "finance@newageproducts.com", "AP / Finance", "Today"),
        ("Demo", "DC Ops", "dc.perris@newageproducts.com", "DC Operations", "Today"),
    ]
    for u in users:
        conn.execute(
            """INSERT OR REPLACE INTO users (first_name, last_name, email, role, last_login)
               VALUES (?, ?, ?, ?, ?)""",
            u,
        )


CARRIER_NAME_NORMALIZE = {
    "Pacific Coastline Drayage Inc.": "Pacific Coastline Drayage",
    "Continental Drayage Solutions, Llc": "Continental Drayage Solutions",
    "Continental Drayage Solutions, Llc.": "Continental Drayage Solutions",
    "Atlantic Container Services, Inc.": "Atlantic Container Services",
}


def _normalize_carrier(name):
    if not name:
        return name
    return CARRIER_NAME_NORMALIZE.get(name, name)


def seed_invoices_and_lines(conn, wb):
    ws_inv = wb["Invoices"]
    h = _header_map(ws_inv, 4)
    for row in _rows_after(ws_inv, 4):
        if not row or not row[h["invoice #"]]:
            continue
        # status is set on the Excel ('AUDIT EXCEPTION' or 'PENDING REVIEW') —
        # we re-map to our 4 canonical statuses after the audit runs.
        # Initially set to 'New'; audit_invoice() will reclassify.
        conn.execute(
            """INSERT OR REPLACE INTO drayage_invoices (invoice_no, carrier_name, invoice_date,
                fb_no, container_no, bol, origin, destination, base_rate, fsc_pct, fsc_amount,
                accessorials_total, grand_total, status, finding_tag, extraction_confidence)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                row[h["invoice #"]], _normalize_carrier(row[h["carrier"]]),
                _val(row[h["invoice date"]]),
                row[h["fb# / load id"]], row[h["container #"]], row[h["bol/mbl #"]],
                row[h["origin"]], row[h["destination"]],
                row[h["linehaul (usd)"]],
                _pct(row[h["fsc %"]]),
                row[h["fsc (usd)"]],
                row[h["accessorials (usd)"]],
                row[h["grand total (usd)"]],
                "New",
                row[h["audit finding"]] if row[h["audit finding"]] != "—" else None,
                1.0,
            ),
        )

    # Lines — clear first to allow re-seed
    conn.execute("DELETE FROM drayage_invoice_lines")
    ws_lines = wb["Invoice_Lines"]
    h = _header_map(ws_lines, 4)
    for row in _rows_after(ws_lines, 4):
        if not row or not row[h["invoice #"]]:
            continue
        conn.execute(
            """INSERT INTO drayage_invoice_lines (invoice_no, line_no, line_type, description,
                qty, rate, amount) VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (row[h["invoice #"]], row[h["line #"]], row[h["type"]], row[h["description"]],
             row[h["qty"]], row[h["rate"]], row[h["amount"]]),
        )


def seed_loads_from_invoices_and_containers(conn):
    """Synthesize a loads row per invoice + one per pre-arrival container."""
    rows = conn.execute(
        """SELECT invoice_no, carrier_name, fb_no, container_no, bol, origin, destination,
            invoice_date FROM drayage_invoices"""
    ).fetchall()
    for r in rows:
        conn.execute(
            """INSERT OR REPLACE INTO loads (fb_no, container_no, carrier_name, bol, origin,
                destination, shipment_date, status, invoice_no)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (r["fb_no"], r["container_no"], r["carrier_name"], r["bol"], r["origin"],
             r["destination"], r["invoice_date"], "Approved", r["invoice_no"]),
        )

    invoiced_containers = set(r["container_no"] for r in rows if r["container_no"])
    container_rows = conn.execute(
        """SELECT number, stage, container_status, days_at_location, lfd, pickup_date,
            steamship_line, vessel, origin_port, us_port FROM containers"""
    ).fetchall()
    for c in container_rows:
        if c["number"] in invoiced_containers:
            continue
        synth_fb = f"PRE-{c['number'][-6:]}"
        status_map = {
            "Awaiting Discharge": "Awaiting Discharge",
            "In Customs": "In Customs",
            "Awaiting Release": "Awaiting Release",
            "Out-Gate Ready": "Ready to Outgate",
            "Delivered": "Delivered" if c["container_status"] == "Empty" else "Completed",
        }
        status = status_map.get(c["stage"], c["stage"])
        conn.execute(
            """INSERT OR REPLACE INTO loads (fb_no, container_no, carrier_name, origin,
                destination, status, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (synth_fb, c["number"], "(pending dispatch)",
             c["us_port"], "NewAge DC", status,
             "Pre-arrival / in-progress load synthesized for visibility."),
        )


def seed_gl_accruals(conn):
    """Synthesize a small GL accrual table for the current period."""
    period = "2026-05"
    accruals = [
        ("5210", "Freight-Inbound", 42500, 38847, -3653, "Open"),
        ("5215", "Duty",             28000, 31200,  3200, "Open"),
        ("5220", "Brokerage",         3400,  3120,  -280, "Open"),
        ("5225", "Demurrage/Detention", 1200, 3010, 1810, "Open"),
    ]
    conn.execute("DELETE FROM gl_accruals WHERE period = ?", (period,))
    for acc, name, accr, actual, var, status in accruals:
        conn.execute(
            """INSERT INTO gl_accruals (period, account_code, account_name, accrued, actual, variance, status)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (period, acc, name, accr, actual, var, status),
        )


if __name__ == "__main__":
    from db import init_db
    init_db()
    seed_all()
    conn = get_conn()
    counts = {}
    for table in ["carriers", "carrier_scorecard", "rate_card", "containers",
                  "purchase_orders", "drayage_invoices", "drayage_invoice_lines",
                  "customs_invoices", "transfer_requests", "p4_transfer_needs", "users",
                  "loads", "terminal_appointments", "carrier_capacity", "per_diem_ladder",
                  "accessorial_rates", "gl_accruals"]:
        c = conn.execute(f"SELECT COUNT(*) AS c FROM {table}").fetchone()["c"]
        counts[table] = c
    conn.close()
    print("Row counts:")
    for k, v in counts.items():
        print(f"  {k}: {v}")
